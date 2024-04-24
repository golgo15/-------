import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import tkinter as tk
from tkinter import ttk
from tkinter import Tk, Listbox
from tkinter import messagebox
import random
import time
import ctypes
import math
import re
from openpyxl.worksheet.hyperlink import Hyperlink
import tkinter.font as fnt
from openpyxl.styles import Font, PatternFill, Border, Side, Color
import datetime

import iconData

#------------------------------------------------
# 定数
#------------------------------------------------
DEBUG_ENABLE = 0    #0:無効、1:有効
DEBUG_MULTIPLE_NUM = 4  # デバッグ用の重複データ数
GLOBAL_KEYS = {"selected_store","yahoo_id"}
SETTING_FILE_NAME = "setting.ini"
PAGE_DATA_NUM = 100

#------------------------------------------------
# グローバル変数
#------------------------------------------------
g_setting_data = {key: None for key in GLOBAL_KEYS}
g_start_page = 1

# 店舗データ
IDX_YAHOO_ID = 0
IDX_STORE_NAME =1
store_data = {
	"西春店":["kaitori_okoku_ya","西春店"]
,	"桑名店":["gzimo46886","桑名店"]
,	"西岐阜店":["slsrh86592","西岐阜店"]
,	"四條畷店":["hhsys91218","四條畷店"]
,	"蟹江店":["aqfei81056","蟹江店"]
,	"春日井19号店":["nkexp48743","春日井19号店"]
,	"大樹寺店":["uqxht92551","大樹寺店"]
,	"京都久世171号店":["cgzyj88481","京都久世171号店"]
,	"津守店":["teaxk49149","津守店"]
,	"堺浜寺２６号店":["jhyzy38060","堺浜寺２６号店"]
,	"大垣２５８店":["qcnnd80919","大垣２５８店"]
,	"鈴鹿白子23号店":["sngbe73253","鈴鹿白子23号店"]
,	"豊川店":["eyemc48423","豊川店"]
,	"多治見19号店":["lslaa56339","多治見19号店"]
,	"八幡店":["muvll59592","八幡店"]
,	"長久手店":["mjeia68505","長久手店"]
,	"守山大森インター店":["ptqaa81694","守山大森インター店"]
,	"東大阪308号店":["zgraj36631","東大阪308号店"]
,	"天理店":["ymqmd24448","天理店"]
,	"金沢鞍月店":["bovgt86014","金沢鞍月店"]
,	"白山福留8号店":["oaana19094","白山福留8号店"]
}


#---------------------------------------------------------
#   関数名:LabelPrintRed()
#   概要:ラベルを表示する
#   引数:[i]:表示オブジェクト
#        [i]:表示する文字列
#   戻り値:なし
#---------------------------------------------------------
def LabelPrintRed(pLabelObj, outStr):
    pLabelObj.config(fg="red", text=outStr, font=("bold"))
    pLabelObj.update()

#---------------------------------------------------------
#   関数名:LabelPrint()
#   概要:ラベルを表示する
#   引数:[i]:表示オブジェクト
#        [i]:表示する文字列
#   戻り値:なし
#---------------------------------------------------------
def LabelPrint(pLabelObj, outStr):
    pLabelObj.config(fg="black", text=outStr, font=("normal"))
    pLabelObj.update()

#---------------------------------------------------------
#   関数名:TxtBoxPrint()
#   概要:テキストボックスを表示する
#   引数:表示する文字列
#   戻り値:なし
#---------------------------------------------------------
def TxtBoxPrint(outStr):
    result_box.insert(tk.END, outStr+'\n')
    result_box.see(tk.END)
    result_box.update()

#---------------------------------------------------------
#   関数名:VisibleProgress()
#   概要:プログレスバーを更新する
#   引数:なし
#   戻り値:なし
#---------------------------------------------------------
def update_progress(add_val):
    progress.configure(value=add_val)
    progress.update()

#---------------------------------------------------------
#   関数名:VisibleProgress()
#   概要:プログレスバーを表示する
#   引数:なし
#   戻り値:なし
#---------------------------------------------------------
def VisibleProgress():
    progress.grid()

#---------------------------------------------------------
#   関数名:HiddenProgress()
#   概要:プログレスバーを隠す
#   引数:なし
#   戻り値:なし
#---------------------------------------------------------
def HiddenProgress():
    progress.grid_remove()


#---------------------------------------------------------
#   関数名:get_yahoo_id()
#   概要:辞書データからYahooIDを取得する
#   引数:なし
#   戻り値:[string]YahooID
#---------------------------------------------------------
def get_yahoo_id():
    selected_store = store_combobox.get()  # プルダウンメニューで選択された店舗名を取得
    if selected_store in store_data:
        yahoo_id = store_data[selected_store][IDX_YAHOO_ID]
        store_name = store_data[selected_store][IDX_STORE_NAME]
        # result_label.config(text=f"{selected_store}のYahoo IDは {yahoo_id} です。")
    else:
        # result_label.config(text=f"{selected_store} に対応する Yahoo ID が見つかりませんでした。")
        yahoo_id = ""
        store_name=""
    return yahoo_id,store_name

#---------------------------------------------------------
#   関数名:read_config()
#   概要:設定ファイルを読み込んで変数に代入する
#   引数:ファイル名
#   戻り値:読み取った内容(key:value)
#---------------------------------------------------------
def read_config(filename):
    config = {}
    try:
        with open(filename, 'r', encoding="utf-8") as file:
            for line in file:
                line = line.strip()  # 改行を除去
                if line.startswith('#') or '=' not in line:
                    continue  # コメント行またはイコールが含まれない行はスキップ
                key, value = line.split('=', 1)  # イコールで行を分割
                if key.strip() in GLOBAL_KEYS:
                    config[key.strip()] = value.strip()  # キーと値を格納
                else:
                    TxtBoxPrint(f"Invalid key: {key.strip()}. Skipping.")
    except FileNotFoundError:
        # 設定ファイル新規作成
        write_config(SETTING_FILE_NAME)

    return config

#---------------------------------------------------------
#   関数名:update_config()
#   概要:設定ファイルを更新する
#   引数:ファイル名
#   戻り値:読み取った内容(key:value)
#---------------------------------------------------------
def update_config(filename, key, new_value):
    updated_lines = []
    found_key = False  # 新しいキーが見つかったかどうかを示すフラグ
    try:
        with open(filename, 'r', encoding="utf-8") as file:
            for line in file:
                if line.strip().startswith(key + "="):
                    line = f"{key}={new_value}\n"  # 新しい値で行を更新
                    found_key = True
                updated_lines.append(line)
    except FileNotFoundError:
        print("ファイル無し")

    # 新しいキーが見つからなかった場合、ファイルの末尾に追加
    if not found_key:
        updated_lines.append(f"{key}={new_value}\n")

    with open(filename, 'w', encoding="utf-8") as file:
        file.writelines(updated_lines)

#---------------------------------------------------------
#   関数名:write_config()
#   概要:設定ファイルを更新する
#   引数:ファイル名
#   戻り値:読み取った内容(key:value)
#---------------------------------------------------------
def write_config(filename):
    # globalsにあるキーを取得して設定ファイルを更新する
    for key in g_setting_data:
        value = g_setting_data[key]  # キーに対応する値を取得
        update_config(filename, key, str(value))  # 設定ファイルを更新

    # TxtBoxPrint("設定ファイルが更新されました。")

#---------------------------------------------------------
#   関数名:DBG_DupDataSet()
#   概要:デバッグ用にdiff_num個の重複データを終端にセットする
#---------------------------------------------------------
def DBG_DupDataSet( all_data, diff_num ):
    # @@debug
    for i in range(0,diff_num):
        auction_id = all_data[i][0]
        auction_id = "4" + auction_id[1:]    #オークションIDは重複しないので先頭を別の文字に変更しておく
        auction_title = all_data[i][1]
        all_data.append((auction_id, auction_title))


#---------------------------------------------------------
#   関数名:GetRandomTime()
#   概要:指定範囲内での乱数（小数第１位まで）を取得する
#---------------------------------------------------------
def GetRandomTime( startNo, EndNo ):
    # 現在の時間をシードとして設定
    seed = int(time.time())

    # 1.0から3.0までの乱数を生成
    random.seed(seed)
    random_number = random.uniform(startNo, EndNo)

    # 小数を第1位までに制限
    random_time = round(random_number, 1)
    # print(f"待機時間: {random_time:.2f}秒")

    return random_time

#---------------------------------------------------------
#   関数名:AnalizeResponse()
#   概要:取得したhtmlを分析する
#   引数:なし
#   戻り値:次ページの有無（0=次ページなし、1=次ページあり）
#---------------------------------------------------------
def AnalizeResponse(response, all_data):
    root.update()

    if response.status_code == 200:
        html = response.text
        soup = BeautifulSoup(html, "html.parser")

        # セレクターを使って要素を選択(出品数)
        selected_elements = soup.select("#allContents > div.gv-l-wrapper.gv-l-wrapper--pc.gv-l-wrapper--liquid > main > div > div.gv-l-contentHeader > div.Tab > ul > li.Tab__item.Tab__item--current > span > span")
        goodsNum = selected_elements[0].text
        LabelPrint(totalNum_label, f"出品数:{goodsNum}")

        # プログレスバーの初期化
        progress["value"] = 0
        str_total = goodsNum[:-1]
        str_total = re.sub(r',', '', str_total)
        int_total = int(str_total)
        endPrgCount = math.floor(int_total / PAGE_DATA_NUM) + 1

        # プログレスバーの更新を開始
        # root.after(1000, update_progress, 100/endPrgCount)  # 最初の更新を1秒後に実行
        update_progress(int((g_start_page+PAGE_DATA_NUM)/PAGE_DATA_NUM/endPrgCount*100))

        # data-auction-idとdata-auction-titleを持つ要素を取得
        items = soup.find_all(lambda tag: tag.has_attr('data-auction-id') and tag.has_attr('data-auction-title'))

        # データをall_dataに追加
        for i in range(0, len(items), 2):  # 2つずつ取得するように変更
            auction_id = items[i]['data-auction-id']
            auction_title = items[i]['data-auction-title']
            all_data.append((auction_id, auction_title))

        # 指定されたセレクターにマッチする要素を取得し、Pager__link--disableが含まれているかチェック
        next_link = soup.select_one("#allContents > div.gv-l-wrapper.gv-l-wrapper--pc.gv-l-wrapper--liquid > main > div > div.gv-l-contentBody > div.gv-l-main > div.Pager > ul > li.Pager__list.Pager__list--next > span.Pager__link--disable")
        if next_link:
            TxtBoxPrint("データ取得完了")
            return 0

        # 次のページへ
        return 1
    else:
        TxtBoxPrint("Failed to retrieve data from Yahoo Auctions")
        return 0

#---------------------------------------------------------
#   関数名:DuplicateCheck()
#   概要:重複チェックを行う
#   引数:[i]all_data        チェック対象データ
#        [o]unique_data     チェック済み（重複無し）データ
#        [o]multiple_data   チェック済み（重複）データ
#   戻り値:なし
#---------------------------------------------------------
def DuplicateCheck(all_data, unique_data, multiple_data):
    seen_titles = set()
    seen_ids = set()
    for auction_id, auction_title in all_data:
        bDuplicated = False
        # ヤフオク仕様が不明なため暫定対応
        # オークションIDが重複するという情報が収集されることがある.
        # ありえないはずだが発生するので、その場合は重複出品ではないと判断する.
        if auction_id not in seen_ids:
            # 重複チェック
            if auction_title not in seen_titles:
                bDuplicated = False  # 重複無し
            else:
                if auction_title in [item[1] for item in unique_data]:  # 重複相手を探す
                    bDuplicated = True  # 重複あり
                else:
                    bDuplicated = False  # 重複無し
        # DBG（オークションID重複検出時の調査用）
        # else:
        #     TxtBoxPrint('\tDBG ---以下、気にしないでください---')
        #     TxtBoxPrint('\tDBG オークションID重複検出')
        #     TxtBoxPrint(f"\tDBG オークションID => {auction_id}")
        #     TxtBoxPrint('\tDBG ---ここまで、気にしないでください---')

        # 重複チェック後のデータセット
        if bDuplicated == True: # 重複あり
            duplicate_partner_id = [item[0] for item in unique_data if item[1] == auction_title][0]
            multiple_data.append((auction_id, auction_title, duplicate_partner_id))
        else:   # 重複なし
            seen_titles.add(auction_title)
            seen_ids.add(auction_id)
            unique_data.append((auction_id, auction_title))




#---------------------------------------------------------
#   関数名:ExportExcelSheet()
#   概要:指定のエクセルシートに指定データを書き込む
#   引数:[i]unique_data     チェック済み（重複無し）データ
#       
#   戻り値:なし
#---------------------------------------------------------
def ExportExcelSheet( ws, unique_data ):
    ws.append(["Auction ID", "Auction Title", "ITCode"])
    for auction_id, auction_title in unique_data:
        ws.append([auction_id, auction_title, auction_title[-12:]])

#---------------------------------------------------------
#   関数名:OutputResultDuplicate()
#   概要:重複あり時の結果を出力する
#   引数:[i]all_data        チェック対象データ
#        [i]unique_data     チェック済み（重複無し）データ
#        [i]multiple_data   チェック済み（重複）データ
#   戻り値:なし
#---------------------------------------------------------
def OutputResultDuplicate( l_work_selStore, all_data, unique_data, multiple_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "出品リスト（ストクリ）"
    ExportExcelSheet( ws, unique_data )
    # 現在の年月日時分を取得
    current_datetime = datetime.datetime.now()
    # 文字列にフォーマットして組み込む
    formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S")

    l_row_cnt = 1
    url_base = "https://page.auctions.yahoo.co.jp/jp/auction/"
    ws_multi = wb.create_sheet("重複出品データ")
    ws_multi.append(["Auction ID(1)", "Auction ID(2)", "Auction Title", "ITCode", "URL<Auction ID(1)>", "URL<Auction ID(2)>"])
    for auction_id, auction_title, duplicate_partner_id in multiple_data:
        url1 = url_base + auction_id
        url2 = url_base + duplicate_partner_id
        ws_multi.append([auction_id, duplicate_partner_id, auction_title, auction_title[-12:], url1, url2])

        ws_multi.cell(row=l_row_cnt+1, column=5).value = f'=HYPERLINK("{url1}", "重複1商品へのリンク")'
        ws_multi.cell(row=l_row_cnt+1, column=6).value = f'=HYPERLINK("{url2}", "重複2商品へのリンク")'
        
        # セルにスタイルを適用
        font = Font(color="0000FF", u='single')  # 赤色のテキスト
        ws_multi.cell(row=l_row_cnt+1, column=5).font = font
        ws_multi.cell(row=l_row_cnt+1, column=6).font = font
        l_row_cnt += 1

    filename = f"{formatted_datetime}_{l_work_selStore}_ヤフオク出品リスト_重複あり.xlsx"
    # filename = "ヤフオク出品リスト_重複あり.xlsx"
    wb.save(filename)
    
    # tkを使うとファイルが大きくなるのでプロンプトで表示に変更
    TxtBoxPrint("重複出品がありましたのでファイルに保存しました。\n重複出品データシートの商品で、ReCOREと紐づいていない商品をストクリにて削除して下さい。")
    TxtBoxPrint(f"ストクリの出品リストを実行ファイルの場所に出力します。\nファイル名 => {filename}")
    TxtBoxPrint("終了します")

    row_count = ws_multi.max_row-1  # ヘッダ行の分-1しておく
    LabelPrintRed(result_label, f"重複:{row_count}件")

#---------------------------------------------------------
#   関数名:OutputResultNoDuplicate()
#   概要:重複なし時の結果を出力する
#   引数:[i]all_data        チェック対象データ
#        [i]unique_data     チェック済み（重複無し）データ
#        [i]multiple_data   チェック済み（重複）データ
#   戻り値:なし
#---------------------------------------------------------
def OutputResultNoDuplicate( l_work_selStore, all_data, unique_data, multiple_data):
    # Excelファイルを作成し、データを書き込む
    wb = Workbook()
    ws = wb.active
    ws.title = "出品リスト（ストクリ）"
    # 現在の年月日時分を取得
    current_datetime = datetime.datetime.now()
    # 文字列にフォーマットして組み込む
    formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S")

    ws.append(["Auction ID", "Auction Title", "ITCode"])
    for auction_id, auction_title in all_data:
        ws.append([auction_id, auction_title, auction_title[-12:]])
    filename = f"{formatted_datetime}_{l_work_selStore}_ヤフオク出品リスト_重複なし.xlsx"
    # filename = "ヤフオク出品リスト_重複なし.xlsx"
    wb.save(filename)

    # tkを使うとファイルが大きくなるのでプロンプトで表示に変更
    TxtBoxPrint("重複出品はありませんでした。")
    TxtBoxPrint(f"ストクリの出品リストを実行ファイルの場所に出力します。\nファイル名 => {filename}")
    TxtBoxPrint("終了します。")
    LabelPrint(result_label, "重複:0件")


#---------------------------------------------------------
#   関数名:OutputResult()
#   概要:結果を出力する
#   引数:[i]all_data        チェック対象データ
#        [i]unique_data     チェック済み（重複無し）データ
#        [i]multiple_data   チェック済み（重複）データ
#   戻り値:なし
#---------------------------------------------------------
def OutputResult( l_work_selStore, all_data, unique_data, multiple_data):
    # データが重複している場合は別のエクセルファイルに書き出す
    if len(all_data) != len(unique_data):
        OutputResultDuplicate( l_work_selStore, all_data, unique_data, multiple_data)
    else:
        OutputResultNoDuplicate( l_work_selStore, all_data, unique_data, multiple_data)


#---------------------------------------------------------
#   関数名:main_proc()
#   概要:メインプロシージャ処理
#   引数:なし
#   戻り値:なし
#---------------------------------------------------------
def main_proc():

    if store_combobox.get() == "":
        LabelPrint(totalNum_label, "店舗を選択してください。")
        LabelPrint(result_label, "--")
        return
    
    LabelPrint(totalNum_label, "--")
    LabelPrint(result_label, "--")

    # (l_yahoo_id,l_store_name) = get_yahoo_id()
    # g_setting_data['yahoo_id']=l_yahoo_id
    base_url = "https://auctions.yahoo.co.jp/seller/{}?sid={}&is_postage_mode=1&dest_pref_code=13&b={}&n={}&mode=2"

    # データを保持するリスト
    all_data = []

    # 設定ファイル更新
    write_config(SETTING_FILE_NAME)
    
    l_work_selStore = g_setting_data["selected_store"]
    TxtBoxPrint(f'{l_work_selStore} ヤフオク重複出品チェック開始')
    global g_start_page
    g_start_page = 1
    update_progress(0)
    VisibleProgress()
    while True:
        time.sleep(GetRandomTime(1.0, 3.0))

        url = base_url.format(g_setting_data['yahoo_id'], g_setting_data['yahoo_id'], g_start_page, PAGE_DATA_NUM)
        response = requests.get(url)
        start_time = time.time()
    
        if 1 == AnalizeResponse(response, all_data):
            g_start_page += PAGE_DATA_NUM
        else:
            break

        end_time = time.time()
        # 処理時間を計算
        elapsed_time = end_time - start_time
        # print(f"処理時間: {elapsed_time:.4f}秒")

    if DEBUG_ENABLE == 1:
        DBG_DupDataSet(all_data, DEBUG_MULTIPLE_NUM)  #debug(先頭から第3引数の個数分だけ同一データを追加して重複させる

    # 重複チェックを行う
    unique_data = []    # 重複しないデータを格納するリスト
    multiple_data = []  # 重複しているデータを格納するリスト
    DuplicateCheck( all_data, unique_data, multiple_data)
    OutputResult( l_work_selStore, all_data, unique_data, multiple_data)
    HiddenProgress()

#---------------------------------------------------------
#   関数名:on_select_store_combobox()
#   概要:コンボボックス変更イベントハンドラ
#   引数:なし
#   戻り値:なし
#---------------------------------------------------------
def on_select_store_combobox(event):
    l_store = store_combobox.get()
    g_setting_data['selected_store'] = l_store
    g_setting_data['yahoo_id'] = store_data.get( l_store )[IDX_YAHOO_ID]
    # 設定ファイル更新
    write_config(SETTING_FILE_NAME)


#---------------------------------------------------------
#   関数名:main()
#   概要:メイン処理
#   引数:なし
#   戻り値:なし
#---------------------------------------------------------
# Tkinterウィンドウの作成
root = tk.Tk()
root.title("ストクリ重複チェック")

ctypes.windll.shcore.SetProcessDpiAwareness(True)

sel_store_name = ""


# プルダウンメニューの作成
# f.Font(family="Lucida Console", weight="bold", size=8, slant="italic")
font1 = fnt.Font(size=12)
store_label = tk.Label(root, justify="left", text="店舗名:")
store_label["font"] = font1
store_label.grid(row=0, column=0, rowspan=2, padx=10, pady=5)
store_combobox = ttk.Combobox(root, values=list(store_data.keys()))
store_combobox.grid(row=0, column=1, rowspan=2, padx=10, pady=5)
store_combobox.bind("<<ComboboxSelected>>", on_select_store_combobox)
store_combobox.set(sel_store_name)
store_combobox["font"] = font1

# 複数行のテキストボックスを作成
result_box = tk.Text(root, height=7, width=60)
result_box.grid(row=3, column=0, columnspan=4, padx=10, pady=5)

# 総件数表示用のラベル
totalNum_label = tk.Label(root, text="--")
totalNum_label.grid(row=0, column=2, columnspan=2, padx=0, pady=5)

# 結果表示用のラベル
result_label = tk.Label(root, text="--")
result_label.grid(row=1, column=2, columnspan=2, padx=0, pady=5)

# 設定ファイルを読み込み、設定を取得
config = read_config(SETTING_FILE_NAME)
# 設定情報を変数に代入
for key, value in config.items():
    g_setting_data[key] = value
sel_store_name = config.get('selected_store')

# プルダウンの初期値セット
initial_index = -1
for index, value in enumerate(store_data.values()):
    if value[1] == sel_store_name:
        initial_index = index
        break
if initial_index != -1:
    store_combobox.set(sel_store_name)
else:
    store_combobox.set("店舗を選択して下さい")


# ボタンの作成
fonts = ("", 18)
search_button = tk.Button(root, bg="#FF9999", text="出品データ取得", height=3, width=30, font=fonts, command=main_proc)
search_button.grid(row=2, column=0, columnspan=4, padx=10, pady=5)
search_button.focus_set()

# プログレスバーの作成
progress = ttk.Progressbar(root, orient="horizontal", length=200, mode="determinate")
progress.grid(row=4, column=0, columnspan=4, padx=10, pady=5)
HiddenProgress()

root.tk.call('wm', 'iconphoto', root._w, tk.PhotoImage(data=iconData.data))

# イベントループの開始
root.mainloop()
#---
